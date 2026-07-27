"""Gera a planilha Excel usada como fonte do Power BI.

Decisoes de formato pensadas para o Power BI:

* Uma aba por box da pagina, criada automaticamente.
* Primeira coluna ``Data``, gravada como DATA de verdade (nao texto).
* ``Valor`` e ``Variacao_pct`` gravados como NUMERO (nao texto).
* Dados empilhados: cada nova data entra abaixo, formando serie historica.
* Cada aba vira uma Tabela nomeada do Excel, o que faz o Power BI reconhecer
  o intervalo automaticamente, mesmo quando cresce.
* Uma aba ``_Indice`` relaciona o nome do box com o nome da aba.
"""

from __future__ import annotations

import logging
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)

# Colunas na ordem em que aparecem na aba. A data vem primeiro, como pedido.
COLUNAS = [
    ("Data", 12),
    ("Localidade", 34),
    ("Valor", 14),
    ("Variacao_pct", 14),
    ("Unidade", 12),
    ("Indicador", 34),
    ("Safra", 12),
    ("Fonte", 14),
    ("Data_Coleta", 20),
]

CARACTERES_PROIBIDOS = r"[]:*?/\\"
LIMITE_NOME_ABA = 31

VERDE = "FF1F7A44"
BRANCO = "FFFFFFFF"


def _sem_acento(texto: str) -> str:
    normalizado = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in normalizado if not unicodedata.combining(c))


# Encurtamentos aplicados quando o titulo estoura o limite do Excel.
ABREVIACOES = [
    ("UTILIZACAO DA CAPACIDADE", "UTL CAP"),
    ("UTL. DA CAP.", "UTL CAP"),
    ("CAPACIDADE FRIGORIFICA", "CAP FRIG"),
    ("FRIGORIFICA", "FRIG"),
    ("INDICADOR DE", ""),
    ("NELORE", "NEL"),
    (" DE ", " "),
    (" DA ", " "),
    (" DO ", " "),
]


def _encurta(texto: str) -> str:
    """Reduz o titulo preservando o sentido, em vez de cortar no meio."""
    resultado = texto
    for busca, troca in ABREVIACOES:
        if len(resultado) <= LIMITE_NOME_ABA:
            break
        resultado = resultado.replace(busca, troca)
    return re.sub(r"\s+", " ", resultado).strip(" -")


def nome_de_aba(titulo: str, usados: set[str]) -> str:
    """Converte o titulo do box em um nome de aba valido e unico.

    O Excel limita a 31 caracteres e proibe ``[ ] : * ? / \\``.
    """
    limpo = _sem_acento(titulo).upper()
    for char in CARACTERES_PROIBIDOS:
        limpo = limpo.replace(char, "-")
    limpo = re.sub(r"\s+", " ", limpo).strip() or "INDICADOR"

    if len(limpo) > LIMITE_NOME_ABA:
        limpo = _encurta(limpo)

    candidato = limpo[:LIMITE_NOME_ABA].strip(" -")
    if candidato.upper() not in usados:
        usados.add(candidato.upper())
        return candidato

    # Resolve colisao acrescentando um sufixo numerico.
    for sufixo in range(2, 100):
        marca = f"_{sufixo}"
        base = limpo[: LIMITE_NOME_ABA - len(marca)].strip()
        candidato = f"{base}{marca}"
        if candidato.upper() not in usados:
            usados.add(candidato.upper())
            return candidato

    raise ValueError(f"Nao consegui gerar nome de aba para {titulo!r}")


def _nome_tabela(indice: int) -> str:
    return f"tbl_{indice:03d}"


def _texto_para_data(valor: str | date | None) -> date | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    try:
        return datetime.strptime(str(valor), "%Y-%m-%d").date()
    except ValueError:
        return None


def _texto_para_datahora(valor: str | None) -> datetime | None:
    if not valor:
        return None
    try:
        return datetime.fromisoformat(str(valor))
    except ValueError:
        return None


class GeradorExcel:
    """Reconstroi a planilha inteira a partir do banco."""

    def __init__(self, caminho: str | Path) -> None:
        self.caminho = Path(caminho)
        self.caminho.parent.mkdir(parents=True, exist_ok=True)

    def gera(self, banco, cadeia: str) -> dict[str, str]:
        """Escreve o arquivo e devolve o mapa {titulo do box: nome da aba}."""
        boxes = banco.boxes_da_cadeia(cadeia)
        if not boxes:
            logger.warning("Nenhum dado no banco para a cadeia %r.", cadeia)
            return {}

        planilha = Workbook()
        planilha.remove(planilha.active)

        usados: set[str] = {"_INDICE"}
        mapa: dict[str, str] = {}

        for indice, box in enumerate(boxes, start=1):
            titulo = box["indicador_nome"]
            if box["safra"]:
                titulo = f"{titulo} - {box['safra']}"

            aba = nome_de_aba(titulo, usados)
            mapa[titulo] = aba

            linhas = banco.historico_do_box(cadeia, box["indicador_id"], box["safra"])
            self._escreve_aba(planilha, aba, linhas, indice)

        self._escreve_indice(planilha, mapa, cadeia)
        planilha.save(self.caminho)
        logger.info("Excel gerado: %s (%d abas)", self.caminho, len(mapa))
        return mapa

    # -- construcao das abas ----------------------------------------------

    def _escreve_aba(self, planilha: Workbook, nome_aba: str, linhas, indice: int) -> None:
        aba = planilha.create_sheet(title=nome_aba)

        for coluna, (titulo, largura) in enumerate(COLUNAS, start=1):
            celula = aba.cell(row=1, column=coluna, value=titulo)
            celula.font = Font(bold=True, color=BRANCO)
            celula.fill = PatternFill("solid", fgColor=VERDE)
            celula.alignment = Alignment(horizontal="center", vertical="center")
            aba.column_dimensions[get_column_letter(coluna)].width = largura

        for numero, registro in enumerate(linhas, start=2):
            aba.cell(row=numero, column=1, value=_texto_para_data(registro["data_referencia"]))
            aba.cell(row=numero, column=2, value=registro["localidade"])
            aba.cell(row=numero, column=3, value=registro["valor"])
            aba.cell(row=numero, column=4, value=registro["variacao"])
            aba.cell(row=numero, column=5, value=registro["unidade"])
            aba.cell(row=numero, column=6, value=registro["indicador_nome"])
            aba.cell(row=numero, column=7, value=registro["safra"] or None)
            aba.cell(row=numero, column=8, value=registro["fonte"] or None)
            aba.cell(row=numero, column=9, value=_texto_para_datahora(registro["coletado_em"]))

            aba.cell(row=numero, column=1).number_format = "DD/MM/YYYY"
            aba.cell(row=numero, column=3).number_format = "#,##0.00"
            aba.cell(row=numero, column=4).number_format = "#,##0.00"
            aba.cell(row=numero, column=9).number_format = "DD/MM/YYYY HH:MM"

        ultima_linha = max(aba.max_row, 2)
        referencia = f"A1:{get_column_letter(len(COLUNAS))}{ultima_linha}"
        tabela = Table(displayName=_nome_tabela(indice), ref=referencia)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showRowStripes=True, showColumnStripes=False
        )
        aba.add_table(tabela)
        aba.freeze_panes = "A2"

    def _escreve_indice(self, planilha: Workbook, mapa: dict[str, str], cadeia: str) -> None:
        aba = planilha.create_sheet(title="_Indice", index=0)
        aba["A1"] = "Cadeia"
        aba["B1"] = "Indicador (box do site)"
        aba["C1"] = "Aba correspondente"
        for coluna in ("A1", "B1", "C1"):
            aba[coluna].font = Font(bold=True, color=BRANCO)
            aba[coluna].fill = PatternFill("solid", fgColor=VERDE)

        for numero, (titulo, nome_aba) in enumerate(mapa.items(), start=2):
            aba.cell(row=numero, column=1, value=cadeia)
            aba.cell(row=numero, column=2, value=titulo)
            aba.cell(row=numero, column=3, value=nome_aba)

        aba.column_dimensions["A"].width = 14
        aba.column_dimensions["B"].width = 44
        aba.column_dimensions["C"].width = 34
        aba.freeze_panes = "A2"
